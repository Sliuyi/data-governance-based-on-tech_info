import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["legislation_wrjs_xny"]

wb = Workbook()
ws = wb.active
i = 2
for x in mycol.find():
    document_number = x["document_number"]
    ws.cell(row=i, column=7).value = document_number
    url = x["url"]
    ws.cell(row=i, column=12).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=6).value = name
    type = x["type"]
    ws.cell(row=i, column=8).value = type
    effective_level = x["effective_level"]
    ws.cell(row=i, column=9).value = effective_level
    timeliness = x["timeliness"]
    ws.cell(row=i, column=10).value = timeliness
    release_date = x["release_date"]
    ws.cell(row=i, column=11).value = release_date
    start_date = x["start_date"]
    ws.cell(row=i, column=11).value = start_date
    content_type = x["content_type"]
    ws.cell(row=i, column=11).value = content_type
    i = i+1
wb.save('legislations.xlsx')