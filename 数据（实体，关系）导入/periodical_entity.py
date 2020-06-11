import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["periodical_wrjs_xny"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    periodical_name = x["periodical_name"]
    ws.cell(row=i, column=7).value = periodical_name
    url = x["url"]
    ws.cell(row=i, column=12).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=6).value = name
    periodical_section = x["periodical_section"]
    ws.cell(row=i, column=8).value = periodical_section
    periodical_year_issue = x["periodical_yea_issue"]
    ws.cell(row=i, column=9).value = periodical_year_issue
    fund_project = x["fund_project"]
    ws.cell(row=i, column=10).value = fund_project
    page_number = x["page_number"]
    ws.cell(row=i, column=11).value = page_number
    i = i+1

wb.save('periodical.xlsx')