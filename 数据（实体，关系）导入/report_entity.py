import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["report_wrjs_xny"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    report_type = x["report_type"]
    ws.cell(row=i, column=7).value = report_type
    url = x["url"]
    ws.cell(row=i, column=12).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=6).value = name
    plan_name = x["plan_name"]
    ws.cell(row=i, column=8).value = plan_name
    plan_year = x["plan_year"]
    ws.cell(row=i, column=9).value = plan_year
    id_number = x["id_number"]
    ws.cell(row=i, column=10).value = id_number
    i = i+1

wb.save('reports.xlsx')