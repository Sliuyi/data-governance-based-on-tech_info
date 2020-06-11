import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["achievement_wrjs_xny"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    year = x["public_year"]
    ws.cell(row=i, column=6).value = year
    url = x["url"]
    ws.cell(row=i, column=7).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=8).value = name
    num = x["project_num"]
    ws.cell(row=i, column=9).value = num
    i = i+1

wb.save('achievements.xlsx')