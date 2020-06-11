import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["organs"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    y = x["name"]
    ws.cell(row=i, column=1).value = y
    ws.cell(row=i, column=6).value = y
    i = i+1

wb.save('jgbm.xlsx')