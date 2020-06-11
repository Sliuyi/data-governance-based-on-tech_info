import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["patent_wrjs_xny"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    apply_num = x["apply_num"]
    ws.cell(row=i, column=7).value = apply_num
    url = x["url"]
    ws.cell(row=i, column=12).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=6).value = name
    apply_date = x["apply_date"]
    ws.cell(row=i, column=8).value = apply_date
    public_num = x["public_num"]
    ws.cell(row=i, column=9).value = public_num
    public_date = x["public_date"]
    ws.cell(row=i, column=10).value = public_date
    apply_man = x["apply_man"]
    ws.cell(row=i, column=11).value = apply_man
    abstract = x["abstract"]
    ws.cell(row=i, column=11).value = abstract
    i = i+1

wb.save('patents.xlsx')