import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["conference_wrjs_xny"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    conference_name = x["conference_name"]
    ws.cell(row=i, column=7).value = conference_name
    url = x["url"]
    ws.cell(row=i, column=10).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=6).value = name
    conference_time = x["conference_time"]
    ws.cell(row=i, column=8).value = conference_time
    conference_local = x["conference_local"]
    ws.cell(row=i, column=9).value = conference_local
    i = i+1

wb.save('conferences.xlsx')