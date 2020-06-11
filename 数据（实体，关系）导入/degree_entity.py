import pymongo
from bson.objectid import ObjectId
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myclient=pymongo.MongoClient(host='127.0.0.1', port=27017)   #指定主机和端口号创建客户端
mydb = myclient["tech_info"]
mycol = mydb["degree_wrjs_xny"]

wb = Workbook()
ws = wb.active

i = 2
for x in mycol.find():
    degree = x["degree"]
    ws.cell(row=i, column=11).value = degree
    url = x["url"]
    ws.cell(row=i, column=8).value = url
    name = x["name"]
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=9).value = name
    doi = x["doi"]
    ws.cell(row=i, column=10).value = doi
    time = x["time"]
    ws.cell(row=i, column=7).value = time
    subject = x["subject"]
    ws.cell(row=i, column=6).value = subject
    i = i+1

wb.save('degrees.xlsx')